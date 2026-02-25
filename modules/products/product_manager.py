# -*- coding: utf-8 -*-
"""
Gestionnaire de produits et de stock
"""
from typing import List, Optional, Dict, Any
from datetime import datetime, timedelta
from database.db_manager import db
from core.logger import logger
from core.data_signals import data_signals
import config


class ProductManager:
    """Gestionnaire de produits"""
    
    def create_product(self, name: str, selling_price: float, 
                      purchase_price: float = 0.0, barcode: str = None,
                      name_ar: str = None, description: str = None,
                      category_id: int = None, stock_quantity: int = 0,
                      min_stock_level: int = 10, unit: str = "pièce",
                      expiry_date: str = None, manufacturing_date: str = None,
                      supplier_id: int = None, created_by: int = None,
                      is_tobacco: int = 0, parent_product_id: int = None,
                      packing_quantity: int = 20) -> tuple[bool, str, Optional[int]]:
        """
        Créer un nouveau produit
        
        Args:
            name: Nom du produit
            selling_price: Prix de vente
            purchase_price: Prix d'achat
            barcode: Code-barres
            name_ar: Nom en arabe
            description: Description
            category_id: ID de la catégorie
            stock_quantity: Quantité en stock
            min_stock_level: Niveau minimum de stock
            unit: Unité (pièce, kg, litre, etc.)
            expiry_date: Date d'expiration (YYYY-MM-DD)
            manufacturing_date: Date de fabrication (YYYY-MM-DD)
            supplier_id: ID du fournisseur
            created_by: ID de l'utilisateur créateur
            is_tobacco: Indique si le produit est du tabac (0 ou 1)
            parent_product_id: ID du produit parent (pour les produits groupés)
            packing_quantity: Quantité d'unités dans un emballage
            
        Returns:
            (success, message, product_id)
        """
        try:
            # Si pas de code-barres, en générer un automatiquement
            if not barcode or barcode.strip() == '':
                import random
                timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                random_suffix = random.randint(100, 999)
                barcode = f"AUTO-{timestamp}-{random_suffix}"
            
            # Vérifier si le code-barres existe déjà
            check_query = "SELECT id, is_active FROM products WHERE barcode = ?"
            existing = db.fetch_one(check_query, (barcode,))
            
            if existing:
                product_id, is_active = existing
                if is_active:
                    return False, "Ce code-barres existe déjà", None
                else:
                    # Réactiver le produit
                    update_query = """
                        UPDATE products 
                        SET name = ?, name_ar = ?, description = ?, category_id = ?,
                            purchase_price = ?, selling_price = ?, stock_quantity = ?, min_stock_level = ?,
                            unit = ?, expiry_date = ?, manufacturing_date = ?, supplier_id = ?, is_active = 1,
                            created_by = ?, is_tobacco = ?, parent_product_id = ?, packing_quantity = ?
                        WHERE id = ?
                    """
                    db.execute_update(update_query, (
                        name, name_ar, description, category_id,
                        purchase_price, selling_price, stock_quantity, min_stock_level,
                        unit, expiry_date, manufacturing_date, supplier_id, created_by,
                        is_tobacco, parent_product_id, packing_quantity,
                        product_id
                    ))
                    
                    logger.info(f"Produit réactivé: {name} (ID: {product_id})")
                    data_signals.product_added.emit()
                    data_signals.products_changed.emit()
                    return True, "Produit réactivé avec succès", product_id
            
            # Insérer le produit
            insert_query = """
                INSERT INTO products (
                    barcode, name, name_ar, description, category_id,
                    purchase_price, selling_price, stock_quantity, min_stock_level,
                    unit, expiry_date, manufacturing_date, supplier_id, created_by,
                    is_tobacco, parent_product_id, packing_quantity
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            
            product_id = db.execute_insert(insert_query, (
                barcode, name, name_ar, description, category_id,
                purchase_price, selling_price, stock_quantity, min_stock_level,
                unit, expiry_date, manufacturing_date, supplier_id, created_by,
                is_tobacco, parent_product_id, packing_quantity
            ))
            
            logger.info(f"Produit créé: {name} (ID: {product_id})")
            
            # Vérifier le stock minimum
            if stock_quantity <= min_stock_level:
                logger.log_stock_alert(name, stock_quantity)
            
            data_signals.product_added.emit()
            data_signals.products_changed.emit()
            return True, "Produit créé avec succès", product_id
            
        except Exception as e:
            error_msg = f"Erreur lors de la création du produit: {str(e)}"
            logger.error(error_msg)
            return False, error_msg, None
    
    def update_product(self, product_id: int, **kwargs) -> tuple[bool, str]:
        """
        Mettre à jour un produit
        
        Args:
            product_id: ID du produit
            **kwargs: Champs à mettre à jour
            
        Returns:
            (success, message)
        """
        try:
            # Champs autorisés
            allowed_fields = [
                'barcode', 'name', 'name_ar', 'description', 'category_id',
                'purchase_price', 'selling_price', 'stock_quantity', 'min_stock_level',
                'unit', 'expiry_date', 'manufacturing_date', 'supplier_id',
                'discount_percentage', 'is_on_promotion',
                'is_tobacco', 'parent_product_id', 'packing_quantity'
            ]
            
            # Filtrer les champs
            updates = []
            params = []
            
            for field, value in kwargs.items():
                if field in allowed_fields:
                    updates.append(f"{field} = ?")
                    params.append(value)
            
            if not updates:
                return False, "Aucune modification à effectuer"
            
            params.append(product_id)
            
            # Exécuter la mise à jour
            query = f"UPDATE products SET {', '.join(updates)} WHERE id = ?"
            rows_affected = db.execute_update(query, tuple(params))
            
            if rows_affected > 0:
                logger.info(f"Produit mis à jour: ID {product_id}")
                
                # Vérifier le stock si modifié
                if 'stock_quantity' in kwargs:
                    product = self.get_product(product_id)
                    if product and product['stock_quantity'] <= product['min_stock_level']:
                        logger.log_stock_alert(product['name'], product['stock_quantity'])
                
                data_signals.product_updated.emit()
                data_signals.products_changed.emit()
                return True, "Produit mis à jour avec succès"
            else:
                return False, "Produit introuvable"
                
        except Exception as e:
            error_msg = f"Erreur lors de la mise à jour: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def delete_product(self, product_id: int) -> tuple[bool, str]:
        """
        Supprimer un produit (soft delete)
        
        Args:
            product_id: ID du produit
            
        Returns:
            (success, message)
        """
        try:
            query = "UPDATE products SET is_active = 0 WHERE id = ?"
            rows_affected = db.execute_update(query, (product_id,))
            
            if rows_affected > 0:
                logger.info(f"Produit supprimé: ID {product_id}")
                data_signals.product_deleted.emit()
                data_signals.products_changed.emit()
                return True, "Produit supprimé avec succès"
            else:
                return False, "Produit introuvable"
                
        except Exception as e:
            error_msg = f"Erreur lors de la suppression: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def get_product(self, product_id: int) -> Optional[Dict]:
        """
        Obtenir un produit par son ID
        
        Args:
            product_id: ID du produit
            
        Returns:
            Dictionnaire avec les données du produit ou None
        """
        query = """
            SELECT p.*, c.name as category_name, s.company_name as supplier_name
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.id
            LEFT JOIN suppliers s ON p.supplier_id = s.id
            WHERE p.id = ?
        """
        result = db.fetch_one(query, (product_id,))
        return dict(result) if result else None
    
    def get_product_by_barcode(self, barcode: str) -> Optional[Dict]:
        """
        Obtenir un produit par son code-barres
        
        Args:
            barcode: Code-barres
            
        Returns:
            Dictionnaire avec les données du produit ou None
        """
        query = """
            SELECT p.*, c.name as category_name, s.company_name as supplier_name
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.id
            LEFT JOIN suppliers s ON p.supplier_id = s.id
            WHERE p.barcode = ? AND p.is_active = 1
        """
        result = db.fetch_one(query, (barcode,))
        return dict(result) if result else None
    
    def get_product_by_name(self, name: str) -> Optional[Dict]:
        """
        Obtenir un produit par son nom exact
        
        Args:
            name: Nom du produit
            
        Returns:
            Dictionnaire avec les données du produit ou None
        """
        query = """
            SELECT p.*, c.name as category_name, s.company_name as supplier_name
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.id
            LEFT JOIN suppliers s ON p.supplier_id = s.id
            WHERE p.name = ? AND p.is_active = 1
        """
        result = db.fetch_one(query, (name,))
        return dict(result) if result else None
    
    def search_products(self, search_term: str, category_id: int = None,
                       include_inactive: bool = False) -> List[Dict]:
        """
        Rechercher des produits
        
        Args:
            search_term: Terme de recherche (nom, code-barres)
            category_id: Filtrer par catégorie
            include_inactive: Inclure les produits désactivés
            
        Returns:
            Liste de produits correspondants
        """
        query = """
            SELECT p.*, c.name as category_name
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.id
            WHERE (p.name LIKE ? OR p.name_ar LIKE ? OR p.barcode LIKE ?)
        """
        
        params = [f"%{search_term}%", f"%{search_term}%", f"%{search_term}%"]
        
        if category_id:
            query += " AND p.category_id = ?"
            params.append(category_id)
        
        if not include_inactive:
            query += " AND p.is_active = 1"
        
        query += " ORDER BY p.name LIMIT 100"
        
        results = db.execute_query(query, tuple(params))
        return [dict(row) for row in results]
    
    def get_all_products(self, category_id: int = None, 
                        include_inactive: bool = False,
                        limit: int = None) -> List[Dict]:
        """
        Obtenir tous les produits
        
        Args:
            category_id: Filtrer par catégorie
            include_inactive: Inclure les produits désactivés
            limit: Limiter le nombre de résultats
            
        Returns:
            Liste de produits
        """
        query = """
            SELECT p.*, c.name as category_name
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.id
            WHERE 1=1
        """
        
        params = []
        
        if category_id:
            query += " AND p.category_id = ?"
            params.append(category_id)
        
        if not include_inactive:
            query += " AND p.is_active = 1"
        
        query += " ORDER BY p.name"
        
        if limit:
            query += f" LIMIT {limit}"
        
        results = db.execute_query(query, tuple(params) if params else ())
        return [dict(row) for row in results]
    
    def update_stock(self, product_id: int, quantity_change: int, 
                    reason: str = "adjustment") -> tuple[bool, str]:
        """
        Mettre à jour le stock d'un produit
        
        Args:
            product_id: ID du produit
            quantity_change: Changement de quantité (positif ou négatif)
            reason: Raison du changement
            
        Returns:
            (success, message)
        """
        try:
            # Obtenir le stock actuel
            product = self.get_product(product_id)
            if not product:
                return False, "Produit introuvable"
            
            new_quantity = product['stock_quantity'] + quantity_change
            
            if new_quantity < 0:
                return False, "Stock insuffisant"
            
            # Mettre à jour le stock
            query = "UPDATE products SET stock_quantity = ? WHERE id = ?"
            db.execute_update(query, (new_quantity, product_id))
            
            logger.info(f"Stock mis à jour: {product['name']} - {quantity_change:+d} ({reason})")
            
            # Vérifier le stock minimum
            if new_quantity <= product['min_stock_level']:
                logger.log_stock_alert(product['name'], new_quantity)
            
            data_signals.product_updated.emit()
            data_signals.products_changed.emit()
            return True, f"Stock mis à jour: {new_quantity}"
            
        except Exception as e:
            error_msg = f"Erreur lors de la mise à jour du stock: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def decrease_stock(self, product_id: int, quantity: int) -> tuple[bool, str]:
        """
        Décrémenter le stock (lors d'une vente)
        
        Args:
            product_id: ID du produit
            quantity: Quantité à décrémenter
            
        Returns:
            (success, message)
        """
        return self.update_stock(product_id, -quantity, "sale")
    
    def increase_stock(self, product_id: int, quantity: int) -> tuple[bool, str]:
        """
        Incrémenter le stock (lors d'un réapprovisionnement)
        
        Args:
            product_id: ID du produit
            quantity: Quantité à ajouter
            
        Returns:
            (success, message)
        """
        return self.update_stock(product_id, quantity, "restock")
    
    def get_low_stock_products(self) -> List[Dict]:
        """
        Obtenir les produits avec stock faible
        
        Returns:
            Liste de produits
        """
        query = """
            SELECT p.*, c.name as category_name
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.id
            WHERE p.stock_quantity <= p.min_stock_level 
              AND p.is_active = 1
              AND p.parent_product_id IS NULL
            ORDER BY p.stock_quantity ASC
        """
        results = db.execute_query(query)
        return [dict(row) for row in results]
    
    def get_expiring_products(self, days: int = None) -> List[Dict]:
        """
        Obtenir les produits qui expirent bientôt
        
        Args:
            days: Nombre de jours avant expiration (si None, utilise la config)
            
        Returns:
            Liste de produits
        """
        if days is None:
            try:
                res = db.fetch_one("SELECT setting_value FROM settings WHERE setting_key = 'expiry_warning_days'")
                days = int(res['setting_value']) if res else 7
            except:
                days = 7

        query = """
            SELECT p.*, c.name as category_name,
                   CAST((julianday(p.expiry_date) - julianday('now')) AS INTEGER) as days_until_expiry
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.id
            WHERE p.expiry_date IS NOT NULL 
            AND p.expiry_date > date('now')
            AND p.expiry_date <= date('now', '+' || ? || ' days')
            AND p.is_active = 1
            ORDER BY p.expiry_date ASC
        """
        results = db.execute_query(query, (days,))
        return [dict(row) for row in results]
    
    def get_expired_products(self) -> List[Dict]:
        """
        Obtenir les produits expirés
        
        Returns:
            Liste de produits expirés
        """
        query = """
            SELECT p.*, c.name as category_name
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.id
            WHERE p.expiry_date IS NOT NULL 
              AND p.expiry_date <= date('now')
              AND p.is_active = 1
            ORDER BY p.expiry_date DESC
        """
        results = db.execute_query(query)
        return [dict(row) for row in results]
    
    def set_promotion(self, product_id: int, discount_percentage: float) -> tuple[bool, str]:
        """
        Mettre un produit en promotion
        
        Args:
            product_id: ID du produit
            discount_percentage: Pourcentage de réduction (0-100)
            
        Returns:
            (success, message)
        """
        try:
            if discount_percentage < 0 or discount_percentage > 100:
                return False, "Le pourcentage doit être entre 0 et 100"
            
            is_on_promotion = 1 if discount_percentage > 0 else 0
            
            query = """
                UPDATE products 
                SET discount_percentage = ?, is_on_promotion = ?
                WHERE id = ?
            """
            rows_affected = db.execute_update(query, (discount_percentage, is_on_promotion, product_id))
            
            if rows_affected > 0:
                logger.info(f"Promotion appliquée: Produit ID {product_id} - {discount_percentage}%")
                return True, "Promotion appliquée avec succès"
            else:
                return False, "Produit introuvable"
                
        except Exception as e:
            error_msg = f"Erreur lors de l'application de la promotion: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def get_promoted_products(self) -> List[Dict]:
        """
        Obtenir les produits en promotion
        
        Returns:
            Liste de produits en promotion
        """
        query = """
            SELECT p.*, c.name as category_name,
                   ROUND(p.selling_price * (1 - p.discount_percentage / 100.0), 2) as discounted_price
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.id
            WHERE p.is_on_promotion = 1 AND p.is_active = 1
            ORDER BY p.discount_percentage DESC
        """
        results = db.execute_query(query)
        return [dict(row) for row in results]
    
    def get_price_history(self, product_id: int) -> List[Dict]:
        """
        Obtenir l'historique des prix d'un produit
        
        Args:
            product_id: ID du produit
            
        Returns:
            Liste des changements de prix
        """
        query = """
            SELECT ph.*, u.full_name as changed_by_name
            FROM price_history ph
            LEFT JOIN users u ON ph.changed_by = u.id
            WHERE ph.product_id = ?
            ORDER BY ph.changed_at DESC
        """
        results = db.execute_query(query, (product_id,))
        return [dict(row) for row in results]
    
    def get_product_stats(self) -> Dict[str, Any]:
        """
        Obtenir des statistiques sur les produits
        
        Returns:
            Dictionnaire avec les statistiques
        """
        stats = {}
        
        # Nombre total de produits
        query = "SELECT COUNT(*) as count FROM products WHERE is_active = 1"
        result = db.fetch_one(query)
        stats['total_products'] = result['count'] if result else 0
        
        # Valeur totale du stock
        query = "SELECT SUM(stock_quantity * purchase_price) as value FROM products WHERE is_active = 1"
        result = db.fetch_one(query)
        stats['total_stock_value'] = round(result['value'], 2) if result and result['value'] else 0.0
        
        # Produits en stock faible
        stats['low_stock_count'] = len(self.get_low_stock_products())
        
        # Produits expirant bientôt
        stats['expiring_soon_count'] = len(self.get_expiring_products())
        
        # Produits expirés
        stats['expired_count'] = len(self.get_expired_products())
        
        # Produits en promotion
        stats['promoted_count'] = len(self.get_promoted_products())
        
        return stats


    def import_products_from_excel(self, file_path: str, created_by: int) -> tuple[bool, Dict]:
        """
        Importer des produits depuis un fichier Excel
        
        Args:
            file_path: Chemin du fichier Excel
            created_by: ID de l'utilisateur
            
        Returns:
            (success, stats)
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            stats = {
                'total': 0,
                'success': 0,
                'errors': 0,
                'duplicates': 0,
                'error_details': []
            }
            
            # En-têtes attendues (flexibles)
            # colonnes: code, nom, prix_achat, prix_vente, stock, min_stock
            
            headers = [cell.value for cell in ws[1]]
            
            # Mapping basique des colonnes
            col_map = {}
            for i, h in enumerate(headers):
                if not h: continue
                h = str(h).lower().strip()
                if 'code' in h or 'barcode' in h: col_map['barcode'] = i
                elif 'nom' in h or 'product' in h: col_map['name'] = i
                elif 'achat' in h or 'purchase' in h: col_map['purchase'] = i
                elif 'vente' in h or 'price' in h or 'selling' in h: col_map['selling'] = i
                elif 'stock' in h and 'min' not in h: col_map['stock'] = i
                elif 'min' in h: col_map['min'] = i
            
            if 'name' not in col_map or 'selling' not in col_map:
                return False, {'error': "Colonnes requises manquantes: 'Nom' et 'Prix Vente'"}
                
            db.begin_transaction()
            
            try:
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row[col_map['name']]: continue # Ignorer lignes vides
                    
                    stats['total'] += 1
                    
                    try:
                        name = str(row[col_map['name']])
                        selling_price = float(row[col_map['selling']])
                        
                        barcode = str(row[col_map['barcode']]) if 'barcode' in col_map and row[col_map['barcode']] else None
                        purchase_price = float(row[col_map['purchase']]) if 'purchase' in col_map and row[col_map['purchase']] else 0.0
                        stock = int(row[col_map['stock']]) if 'stock' in col_map and row[col_map['stock']] else 0
                        min_stock = int(row[col_map['min']]) if 'min' in col_map and row[col_map['min']] else 10
                        
                        # Créer le produit
                        success, msg, _ = self.create_product(
                            name=name,
                            selling_price=selling_price,
                            purchase_price=purchase_price,
                            barcode=barcode,
                            stock_quantity=stock,
                            min_stock_level=min_stock,
                            created_by=created_by
                        )
                        
                        if success:
                            stats['success'] += 1
                        else:
                            if "existe déjà" in msg:
                                stats['duplicates'] += 1
                            else:
                                stats['errors'] += 1
                                stats['error_details'].append(f"Ligne {row_idx}: {msg}")
                                
                    except Exception as e:
                        stats['errors'] += 1
                        stats['error_details'].append(f"Ligne {row_idx}: {str(e)}")
                
                db.commit()
                return True, stats
                
            except Exception as e:
                db.rollback()
                raise e
                
        except Exception as e:
            logger.error(f"Erreur import Excel: {e}")
            return False, {'error': str(e)}


# Instance globale
product_manager = ProductManager()
