# -*- coding: utf-8 -*-
"""
Interface des paramètres
"""
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                             QPushButton, QLineEdit, QTableWidget, QTableWidgetItem,
                             QComboBox, QFrame, QMessageBox, QHeaderView, QTabWidget,
                             QFormLayout, QGroupBox, QCheckBox, QSpinBox, QFileDialog)
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QFont, QColor, QPixmap
from core.auth import auth_manager
from core.logger import logger
from database.db_manager import db
import config
import openpyxl
import os
from ui.permission_dialog import PermissionDialog

from core.i18n import i18n_manager

class SettingsPage(QWidget):
    """Page de configuration"""
    
    # Signal pour changement de thème
    theme_changed = pyqtSignal(bool)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.init_ui()
        # Connect to language change
        i18n_manager.language_changed.connect(self.update_ui_text)
        
    def init_ui(self):
        # Create a main layout if it doesn't exist
        if not self.layout():
            layout = QVBoxLayout()
            layout.setContentsMargins(0, 0, 0, 0)
            self.setLayout(layout)
        
        # Create fresh container
        self.container = QWidget()
        self.layout().addWidget(self.container)
        
        # Build UI inside container
        self.build_ui_content(self.container)
        self.load_users() # Reload users after build

    def update_ui_text(self):
        """Mettre à jour les textes de l'interface"""
        # Remove old container
        if hasattr(self, 'container'):
            if self.layout():
                self.layout().removeWidget(self.container)
            self.container.deleteLater()
        
        # Create new container
        self.container = QWidget()
        self.layout().addWidget(self.container)
        
        # Update layout direction
        if i18n_manager.is_rtl():
             self.setLayoutDirection(Qt.RightToLeft)
        else:
             self.setLayoutDirection(Qt.LeftToRight)
        
        # Rebuild UI
        self.build_ui_content(self.container)
        self.load_users()

    def build_ui_content(self, parent_widget):
        _ = i18n_manager.get
        
        layout = QVBoxLayout(parent_widget)
        
        from ui._styles import (header_style, HEADER_TITLE_STYLE, HEADER_SUBTITLE_STYLE,
                                TAB_WIDGET_STYLE, GROUP_BOX_STYLE, TABLE_STYLE,
                                FORM_INPUT_STYLE, PRIMARY_BTN, DANGER_BTN, SECONDARY_BTN,
                                GREEN_BTN, PURPLE_BTN)
        
        # En-tête avec gradient
        header_frame = QFrame()
        header_frame.setStyleSheet(header_style("#475569", "#334155"))
        header_layout = QHBoxLayout(header_frame)
        
        title_layout = QVBoxLayout()
        header = QLabel(_('settings_title'))
        header.setStyleSheet(HEADER_TITLE_STYLE)
        title_layout.addWidget(header)
        
        subtitle = QLabel(_('settings_subtitle'))
        subtitle.setStyleSheet(HEADER_SUBTITLE_STYLE)
        title_layout.addWidget(subtitle)
        
        header_layout.addLayout(title_layout)
        header_layout.addStretch()
        layout.addWidget(header_frame)
        
        # Onglets stylisés
        tabs = QTabWidget()
        tabs.setStyleSheet(TAB_WIDGET_STYLE)
        
        is_admin = auth_manager.is_admin()
        
        if is_admin:
            # Onglet Utilisateurs
            self.users_tab = self.create_users_tab()
            tabs.addTab(self.users_tab, _('tab_users'))
            
            # Onglet Données (Export)
            self.data_tab = self.create_data_tab()
            tabs.addTab(self.data_tab, _('tab_data'))

            # Onglet Magasin
            self.store_tab = self.create_store_tab()
            tabs.addTab(self.store_tab, _('tab_store'))
            
            # Onglet Sécurité
            self.security_tab = self.create_security_tab()
            tabs.addTab(self.security_tab, "Sécurité")
        
        # Onglet Tutoriel (Pour tous)
        self.tutorial_tab = self.create_tutorial_tab()
        tabs.addTab(self.tutorial_tab, _('tab_tutorial'))
        
        # Onglet À propos (Pour tous)
        self.about_tab = self.create_about_tab()
        tabs.addTab(self.about_tab, _('tab_about'))
        
        layout.addWidget(tabs)

    def create_security_tab(self):
        """Onglet de sécurité"""
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Safe Security Group
        safe_group = QGroupBox("Sécurité Coffre")
        form = QFormLayout()
        
        # Status indicator
        self.safe_password_status = QLabel()
        try:
            res = db.fetch_one("SELECT setting_value FROM settings WHERE setting_key = 'safe_password'")
            has_password = res and res['setting_value']
            if has_password:
                self.safe_password_status.setText("✅ Mot de passe défini")
                self.safe_password_status.setStyleSheet("color: #27ae60; font-weight: bold;")
            else:
                self.safe_password_status.setText("⚠️ Aucun mot de passe (accès libre)")
                self.safe_password_status.setStyleSheet("color: #e67e22; font-weight: bold;")
        except Exception as e:
            logger.error(f"Error loading safe password status: {e}")
            self.safe_password_status.setText("❓ Erreur de chargement")
        
        form.addRow("État actuel:", self.safe_password_status)
        
        # New password field (always empty - user types new password)
        self.safe_password_edit = QLineEdit()
        self.safe_password_edit.setEchoMode(QLineEdit.Password)
        self.safe_password_edit.setPlaceholderText("Entrez un nouveau mot de passe (vide = désactiver)")
        
        form.addRow("Nouveau mot de passe:", self.safe_password_edit)
        
        # Buttons
        btn_layout = QHBoxLayout()
        
        save_btn = QPushButton("💾 Enregistrer")
        save_btn.clicked.connect(self.save_safe_security)
        save_btn.setStyleSheet("background-color: #27ae60; color: white; padding: 8px;")
        
        clear_btn = QPushButton("🗑️ Supprimer le mot de passe")
        clear_btn.clicked.connect(self.clear_safe_password)
        clear_btn.setStyleSheet("background-color: #e74c3c; color: white; padding: 8px;")
        
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(clear_btn)
        form.addRow(btn_layout)
        
        safe_group.setLayout(form)
        layout.addWidget(safe_group)
        layout.addStretch()
        
        tab.setLayout(layout)
        return tab
    
    def clear_safe_password(self):
        """Supprimer le mot de passe du coffre"""
        reply = QMessageBox.question(self, "Confirmation", 
            "Voulez-vous vraiment supprimer le mot de passe du coffre?\nL'accès sera libre.",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            try:
                db.execute_update("DELETE FROM settings WHERE setting_key = 'safe_password'")
                self.safe_password_status.setText("⚠️ Aucun mot de passe (accès libre)")
                self.safe_password_status.setStyleSheet("color: #e67e22; font-weight: bold;")
                self.safe_password_edit.clear()
                QMessageBox.information(self, "Succès", "Mot de passe supprimé")
                logger.info("Safe password cleared")
            except Exception as e:
                QMessageBox.critical(self, "Erreur", str(e))


    def save_safe_security(self):
        """Enregistrer le mot de passe du coffre"""
        password = self.safe_password_edit.text().strip()
        
        if not password:
            # User submitted empty - same as clear
            reply = QMessageBox.question(self, "Confirmation", 
                "Le champ est vide. Voulez-vous supprimer le mot de passe?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                self.clear_safe_password()
            return
        
        try:
            db.execute_update(
                "INSERT OR REPLACE INTO settings (setting_key, setting_value) VALUES ('safe_password', ?)", 
                (password,)
            )
            # Update status
            self.safe_password_status.setText("✅ Mot de passe défini")
            self.safe_password_status.setStyleSheet("color: #27ae60; font-weight: bold;")
            self.safe_password_edit.clear()
            
            QMessageBox.information(self, "Succès", "Mot de passe coffre mis à jour")
            logger.info("Safe password updated")
        except Exception as e:
            QMessageBox.critical(self, "Erreur", str(e))
            logger.error(f"Error saving safe password: {e}")



    def create_data_tab(self):
        """Onglet de gestion des données"""
        _ = i18n_manager.get
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Section Configuration Sauvegarde Auto
        backup_config_group = QGroupBox(_('group_backup_config'))
        backup_form = QFormLayout()
        
        self.auto_backup_check = QCheckBox(_('check_auto_backup'))
        self.auto_backup_check.setChecked(True) # Valeur par défaut
        
        self.backup_interval_spin = QSpinBox()
        self.backup_interval_spin.setRange(1, 48)
        self.backup_interval_spin.setSuffix(_('suffix_hours'))
        self.backup_interval_spin.setValue(config.BACKUP_CONFIG.get("backup_interval_hours", 5))
        
        # Charger les valeurs depuis la DB
        try:
            # Check enabled
            res_enabled = db.fetch_one("SELECT setting_value FROM settings WHERE setting_key = 'auto_backup_enabled'")
            if res_enabled:
                self.auto_backup_check.setChecked(res_enabled['setting_value'] == '1')
            else:
                self.auto_backup_check.setChecked(config.BACKUP_CONFIG.get("auto_backup", True))
                
            # Check interval
            res_interval = db.fetch_one("SELECT setting_value FROM settings WHERE setting_key = 'backup_interval_hours'")
            if res_interval:
                self.backup_interval_spin.setValue(int(res_interval['setting_value']))
        except Exception as e:
            logger.error(f"Erreur chargement config backup: {e}")
            
        backup_form.addRow(self.auto_backup_check)
        backup_form.addRow(_('label_interval'), self.backup_interval_spin)
        
        save_backup_btn = QPushButton(_('btn_save_config'))
        save_backup_btn.clicked.connect(self.save_backup_config)
        save_backup_btn.setStyleSheet("background-color: #34495e; color: white;")
        backup_form.addRow(save_backup_btn)
        
        backup_config_group.setLayout(backup_form)
        layout.addWidget(backup_config_group)
        
        # Section Export
        export_group = QGroupBox(_('group_export'))
        export_form = QFormLayout()
        
        export_info = QLabel(_('label_export_info'))
        export_info.setStyleSheet("color: gray;")
        export_info.setWordWrap(True)
        export_form.addRow(export_info)
        
        export_btn = QPushButton(_('btn_create_backup'))
        export_btn.setStyleSheet("background-color: #27ae60; color: white; padding: 12px; font-weight: bold; font-size: 14px;")
        export_btn.clicked.connect(self.export_data)
        export_form.addRow(export_btn)
        
        export_group.setLayout(export_form)
        layout.addWidget(export_group)
        
        # Section Import
        import_group = QGroupBox(_('group_import'))
        import_form = QFormLayout()
        
        import_info = QLabel(_('label_import_info'))
        import_info.setStyleSheet("color: gray;")
        import_info.setWordWrap(True)
        import_form.addRow(import_info)
        
        import_btn = QPushButton(_('btn_restore_backup'))
        import_btn.setStyleSheet("background-color: #3498db; color: white; padding: 12px; font-weight: bold; font-size: 14px;")
        import_btn.clicked.connect(self.import_data)
        import_form.addRow(import_btn)
        
        import_group.setLayout(import_form)
        layout.addWidget(import_group)
        
        # Section Gestionnaire de Sauvegardes
        backup_mgr_group = QGroupBox("📁 Gestionnaire de Sauvegardes")
        backup_mgr_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #e2e8f0;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                color: #3498db;
            }
        """)
        backup_mgr_layout = QVBoxLayout()
        
        # Header with total size info
        mgr_header = QHBoxLayout()
        self.backup_total_label = QLabel("Chargement...")
        self.backup_total_label.setStyleSheet("color: #6b7280; font-size: 13px;")
        mgr_header.addWidget(self.backup_total_label)
        mgr_header.addStretch()
        
        refresh_backups_btn = QPushButton("🔄 Rafraîchir")
        refresh_backups_btn.setStyleSheet("background-color: #3498db; color: white; padding: 6px 15px; border-radius: 5px; font-weight: bold;")
        refresh_backups_btn.clicked.connect(self.load_backups_table)
        mgr_header.addWidget(refresh_backups_btn)
        
        delete_all_btn = QPushButton("🗑️ Tout supprimer")
        delete_all_btn.setStyleSheet("background-color: #e74c3c; color: white; padding: 6px 15px; border-radius: 5px; font-weight: bold;")
        delete_all_btn.clicked.connect(self.delete_all_backups)
        mgr_header.addWidget(delete_all_btn)
        
        backup_mgr_layout.addLayout(mgr_header)
        
        # Backups table
        self.backups_table = QTableWidget()
        self.backups_table.setColumnCount(5)
        self.backups_table.setHorizontalHeaderLabels(["Nom", "Type", "Taille", "Date", "Action"])
        self.backups_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.backups_table.setColumnWidth(1, 120)
        self.backups_table.setColumnWidth(2, 100)
        self.backups_table.setColumnWidth(3, 150)
        self.backups_table.setColumnWidth(4, 100)
        self.backups_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.backups_table.setAlternatingRowColors(True)
        self.backups_table.verticalHeader().setDefaultSectionSize(40)
        self.backups_table.setMaximumHeight(300)
        self.backups_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #e5e7eb;
                border-radius: 6px;
                background-color: white;
                font-size: 12px;
            }
            QTableWidget::item { padding: 5px; }
            QTableWidget::item:alternate { background-color: #f9fafb; }
            QHeaderView::section {
                background-color: #f1f5f9;
                padding: 8px;
                border: none;
                border-bottom: 2px solid #e5e7eb;
                font-weight: bold;
                color: #4b5563;
            }
        """)
        backup_mgr_layout.addWidget(self.backups_table)
        
        backup_mgr_group.setLayout(backup_mgr_layout)
        layout.addWidget(backup_mgr_group)
        
        # Load backups on creation
        self.load_backups_table()
        # Section Réinitialisation (DANGER)
        reset_group = QGroupBox(_('group_reset'))
        reset_group.setStyleSheet("""
            QGroupBox {
                border: 2px solid #e74c3c;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                color: #e74c3c;
            }
        """)
        reset_form = QFormLayout()
        
        reset_info = QLabel(_('label_reset_info'))
        reset_info.setStyleSheet("color: #e74c3c; font-weight: bold;")
        reset_info.setWordWrap(True)
        reset_form.addRow(reset_info)
        
        reset_btn = QPushButton(_('btn_reset_all'))
        reset_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c; 
                color: white; 
                padding: 12px; 
                font-weight: bold; 
                font-size: 14px;
                border: 2px solid #c0392b;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        reset_btn.clicked.connect(self.reset_all_data)
        reset_form.addRow(reset_btn)
        
        reset_group.setLayout(reset_form)
        layout.addWidget(reset_group)
        
        layout.addStretch()
        tab.setLayout(layout)
        return tab

    def save_backup_config(self):
        """Enregistrer la configuration de sauvegarde"""
        _ = i18n_manager.get
        try:
            enabled = '1' if self.auto_backup_check.isChecked() else '0'
            interval = str(self.backup_interval_spin.value())
            
            # Upsert settings
            db.execute_update("INSERT OR REPLACE INTO settings (setting_key, setting_value) VALUES ('auto_backup_enabled', ?)", (enabled,))
            db.execute_update("INSERT OR REPLACE INTO settings (setting_key, setting_value) VALUES ('backup_interval_hours', ?)", (interval,))
            
            QMessageBox.information(self, _('title_success'), _('msg_config_saved'))
            logger.info(f"Config backup mise à jour: Enabled={enabled}, Interval={interval}h")
            
        except Exception as e:
            QMessageBox.critical(self, _('title_error'), f"{_('title_error')}: {e}")
            logger.error(f"Erreur save backup config: {e}")

    def load_backups_table(self):
        """Charger la liste des sauvegardes dans la table"""
        try:
            from core.backup import backup_manager
            
            backups = backup_manager.list_backups()
            total_size = backup_manager.get_total_backup_size()
            
            self.backup_total_label.setText(f"📊 {len(backups)} sauvegardes — Espace total: {total_size}")
            
            self.backups_table.setRowCount(0)
            for backup in backups:
                row = self.backups_table.rowCount()
                self.backups_table.insertRow(row)
                
                self.backups_table.setItem(row, 0, QTableWidgetItem(backup['name']))
                self.backups_table.setItem(row, 1, QTableWidgetItem(backup['type']))
                self.backups_table.setItem(row, 2, QTableWidgetItem(backup['size_str']))
                self.backups_table.setItem(row, 3, QTableWidgetItem(backup['created_str']))
                
                # Delete button
                del_btn = QPushButton("🗑️")
                del_btn.setFixedSize(35, 30)
                del_btn.setCursor(Qt.PointingHandCursor)
                del_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #fef2f2;
                        border: 1px solid #fecaca;
                        border-radius: 5px;
                        color: #dc2626;
                        font-size: 14px;
                    }
                    QPushButton:hover {
                        background-color: #fee2e2;
                    }
                """)
                # Store path in button property
                del_btn.setProperty("backup_path", str(backup['path']))
                del_btn.clicked.connect(self.delete_backup_entry)
                self.backups_table.setCellWidget(row, 4, del_btn)
                
        except Exception as e:
            logger.error(f"Erreur chargement sauvegardes: {e}")

    def delete_backup_entry(self):
        """Supprimer une sauvegarde spécifique"""
        from pathlib import Path
        from core.backup import backup_manager
        
        btn = self.sender()
        if not btn:
            return
        
        backup_path = Path(btn.property("backup_path"))
        
        reply = QMessageBox.question(self, "Confirmation", 
            f"Supprimer la sauvegarde ?\n{backup_path.name}",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            success, msg = backup_manager.delete_backup(backup_path)
            if success:
                self.load_backups_table()
            else:
                QMessageBox.warning(self, "Erreur", msg)

    def delete_all_backups(self):
        """Supprimer toutes les sauvegardes"""
        from core.backup import backup_manager
        
        backups = backup_manager.list_backups()
        if not backups:
            QMessageBox.information(self, "Info", "Aucune sauvegarde à supprimer")
            return
        
        reply = QMessageBox.warning(self, "⚠️ Attention", 
            f"Voulez-vous vraiment supprimer les {len(backups)} sauvegardes ?\nCette action est irréversible !",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            deleted = 0
            for backup in backups:
                success, _ = backup_manager.delete_backup(backup['path'])
                if success:
                    deleted += 1
            
            self.load_backups_table()
            QMessageBox.information(self, "Succès", f"{deleted} sauvegarde(s) supprimée(s)")

    def export_data(self):
        """Exporter les données en Excel (sauvegarde complète)"""
        _ = i18n_manager.get
        try:
            from datetime import datetime
            default_name = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filename, selected_filter = QFileDialog.getSaveFileName(self, _('group_export'), 
                                                    str(config.DATA_DIR / default_name), 
                                                    "Fichiers Excel (*.xlsx)")
            if not filename:
                return
            
            # ... (Rest of logic unchanged, just messages)
            # Re-implementing core logic with minimal changes
            
            wb = openpyxl.Workbook()
            
            # 1. Produits
            ws_prod = wb.active
            ws_prod.title = "Produits"
            # 1. Produits
            ws_prod = wb.active
            ws_prod.title = "Produits"
            # Fixed Query: JOIN with categories table to get category name AND use correct min_stock_level column
            products_query = """
                SELECT p.barcode, p.name, c.name as category_name, 
                       p.purchase_price, p.selling_price, p.stock_quantity, p.min_stock_level 
                FROM products p
                LEFT JOIN categories c ON p.category_id = c.id
            """
            products = db.execute_query(products_query)
            ws_prod.append(["Code-barres", "Nom", "Catégorie", "PA", "PV", "Stock", "Stock Min"])
            for p in products:
                ws_prod.append([
                    p['barcode'], 
                    p['name'], 
                    p['category_name'] or '', # Use category_name alias
                    p['purchase_price'], 
                    p['selling_price'], 
                    p['stock_quantity'], 
                    p['min_stock_level'] or 0
                ])
                
            # 2. Ventes
            ws_sales = wb.create_sheet("Ventes")
            sales = db.execute_query("SELECT sale_number, total_amount, payment_method, sale_date, customer_id, status FROM sales")
            ws_sales.append(["N° Vente", "Montant Total", "Paiement", "Date", "Client ID", "Statut"])
            for s in sales:
                ws_sales.append([s['sale_number'], s['total_amount'], s['payment_method'], s['sale_date'], s['customer_id'] or '', s['status']])

            # 3. Détails Ventes (Fixed: subtotal instead of total)
            ws_items = wb.create_sheet("Details_Ventes")
            items = db.execute_query("SELECT sale_id, product_id, product_name, quantity, unit_price, subtotal FROM sale_items")
            ws_items.append(["ID Vente", "ID Produit", "Nom Produit", "Quantité", "Prix Unitaire", "Sous-Total"])
            for i in items:
                ws_items.append([i['sale_id'], i['product_id'] or '', i['product_name'], i['quantity'], i['unit_price'], i['subtotal']])

            # 4. Clients (More columns)
            ws_cust = wb.create_sheet("Clients")
            customers = db.execute_query("SELECT code, full_name, phone, email, address, credit_limit, current_credit, total_purchases FROM customers")
            ws_cust.append(["Code", "Nom", "Téléphone", "Email", "Adresse", "Limite Crédit", "Dette", "Total Achats"])
            for c in customers:
                ws_cust.append([c['code'], c['full_name'], c['phone'] or '', c['email'] or '', c['address'] or '', c['credit_limit'], c['current_credit'], c['total_purchases']])

            # 5. Fournisseurs (Fixed: company_name instead of name)
            ws_sup = wb.create_sheet("Fournisseurs")
            try:
                suppliers = db.execute_query("SELECT code, company_name, contact_person, phone, email, address, total_debt, total_purchases FROM suppliers")
                ws_sup.append(["Code", "Entreprise", "Contact", "Téléphone", "Email", "Adresse", "Dette", "Total Achats"])
                for sup in suppliers:
                    ws_sup.append([sup['code'] or '', sup['company_name'], sup['contact_person'] or '', sup['phone'] or '', sup['email'] or '', sup['address'] or '', sup['total_debt'], sup['total_purchases']])
            except Exception as e:
                logger.warning(f"Erreur export fournisseurs: {e}")
                ws_sup.append(["Aucune donnée fournisseur"])
            
            # 6. Catégories (NEW)
            ws_cat = wb.create_sheet("Categories")
            try:
                categories = db.execute_query("SELECT id, name, name_ar, description FROM categories WHERE is_active = 1")
                ws_cat.append(["ID", "Nom", "Nom Arabe", "Description"])
                for cat in categories:
                    ws_cat.append([cat['id'], cat['name'], cat['name_ar'] or '', cat['description'] or ''])
            except Exception as e:
                logger.warning(f"Erreur export catégories: {e}")
                ws_cat.append(["Aucune catégorie"])

            # 7. Retours (NEW)
            ws_ret = wb.create_sheet("Retours")
            try:
                returns = db.execute_query("SELECT return_number, original_sale_id, return_amount, refund_method, return_date, reason FROM returns")
                ws_ret.append(["N° Retour", "ID Vente Originale", "Montant", "Méthode Remboursement", "Date", "Raison"])
                for r in returns:
                    ws_ret.append([r['return_number'], r['original_sale_id'], r['return_amount'], r['refund_method'], r['return_date'], r['reason'] or ''])
            except Exception as e:
                logger.warning(f"Erreur export retours: {e}")
                ws_ret.append(["Aucun retour"])
            
            # 8. Détails Retours (NEW)
            ws_ret_items = wb.create_sheet("Details_Retours")
            try:
                ret_items = db.execute_query("SELECT return_id, product_id, quantity_returned, unit_price, subtotal FROM return_items")
                ws_ret_items.append(["ID Retour", "ID Produit", "Qté Retournée", "Prix Unitaire", "Sous-Total"])
                for ri in ret_items:
                    ws_ret_items.append([ri['return_id'], ri['product_id'] or '', ri['quantity_returned'], ri['unit_price'], ri['subtotal']])
            except Exception as e:
                logger.warning(f"Erreur export détails retours: {e}")
                ws_ret_items.append(["Aucun détail retour"])
            
            # 9. Raccourcis POS (NEW)
            ws_short = wb.create_sheet("Raccourcis")
            try:
                shortcuts = db.execute_query("SELECT id, label, product_id, category_id, price, image_path, position FROM pos_shortcuts ORDER BY position")
                ws_short.append(["ID", "Libellé", "ID Produit", "ID Catégorie", "Prix", "Chemin Image", "Position"])
                for sc in shortcuts:
                    ws_short.append([sc['id'], sc['label'], sc['product_id'] or '', sc['category_id'] or '', sc['price'] or '', sc['image_path'] or '', sc['position']])
            except Exception as e:
                logger.warning(f"Erreur export raccourcis: {e}")
                ws_short.append(["Aucun raccourci"])
            
            wb.save(filename)
            logger.info(f"Sauvegarde créée: {filename}")
            QMessageBox.information(self, _('title_success'), _('msg_backup_success').format(filename))
            
        except Exception as e:
            logger.error(f"Erreur export excel: {e}")
            QMessageBox.critical(self, _('title_error'), f"{_('title_error')}: {e}")

    def import_data(self):
        """Importer les données depuis une sauvegarde Excel"""
        _ = i18n_manager.get
        reply = QMessageBox.warning(self, _('title_warning'), 
            _('msg_confirm_import'),
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        if reply != QMessageBox.Yes:
            return
            
        try:
            filename, selected_filter = QFileDialog.getOpenFileName(self, _('group_import'),
                                                     str(config.DATA_DIR),
                                                     "Fichiers Excel (*.xlsx)")
            if not filename:
                return

            # ... (Rest logic matches original import mostly)
            wb = openpyxl.load_workbook(filename)
            imported_counts = {}
            
            # 1. Importer Produits
            if "Produits" in wb.sheetnames:
                ws = wb["Produits"]
                count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # barcode exists
                        try:
                            # Handle Category Lookup/Creation
                            cat_name = row[2] or ''
                            cat_id = None
                            if cat_name:
                                # Find or Create Category
                                cat_res = db.fetch_one("SELECT id FROM categories WHERE name = ?", (cat_name,))
                                if cat_res:
                                    cat_id = cat_res['id']
                                else:
                                    # Create new category if not exists
                                    db.execute_update("INSERT INTO categories (name) VALUES (?)", (cat_name,))
                                    cat_res = db.fetch_one("SELECT id FROM categories WHERE name = ?", (cat_name,))
                                    if cat_res:
                                        cat_id = cat_res['id']

                            db.execute_update("""
                                INSERT OR REPLACE INTO products (barcode, name, category_id, purchase_price, selling_price, stock_quantity, min_stock_level)
                                VALUES (?, ?, ?, ?, ?, ?, ?)
                            """, (row[0], row[1], cat_id, row[3] or 0, row[4] or 0, row[5] or 0, row[6] or 0))
                            count += 1
                        except:
                            pass
                imported_counts["Produits"] = count

            # 2. Importer Clients (Updated format: Code, Nom, Tel, Email, Adresse, Limite, Dette, Total)
            if "Clients" in wb.sheetnames:
                ws = wb["Clients"]
                count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[1]:  # full_name (column 1) exists
                        try:
                            # Handle both old (4 cols) and new (8 cols) format
                            if len([x for x in row if x is not None]) >= 8:
                                # New format with all columns
                                db.execute_update("""
                                    INSERT OR IGNORE INTO customers (code, full_name, phone, email, address, credit_limit, current_credit, total_purchases)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                                """, (row[0] or '', row[1], row[2] or '', row[3] or '', row[4] or '', row[5] or 0, row[6] or 0, row[7] or 0))
                            else:
                                # Old format (Nom, Tel, Dette, Total)
                                db.execute_update("""
                                    INSERT OR IGNORE INTO customers (full_name, phone, current_credit, total_purchases)
                                    VALUES (?, ?, ?, ?)
                                """, (row[0], row[1] or '', row[2] or 0, row[3] or 0))
                            count += 1
                        except:
                            pass
                imported_counts["Clients"] = count

            # 3. Importer Fournisseurs (Updated: Code, Entreprise, Contact, Tel, Email, Adresse, Dette, Total)
            if "Fournisseurs" in wb.sheetnames:
                ws = wb["Fournisseurs"]
                count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[1] and row[1] != "Aucune donnée fournisseur":  # company_name
                        try:
                            db.execute_update("""
                                INSERT OR IGNORE INTO suppliers (code, company_name, contact_person, phone, email, address, total_debt, total_purchases)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                            """, (row[0] or '', row[1], row[2] or '', row[3] or '', row[4] or '', row[5] or '', row[6] or 0, row[7] or 0))
                            count += 1
                        except:
                            pass
                imported_counts["Fournisseurs"] = count
            
            # 4. Importer Catégories (NEW)
            if "Categories" in wb.sheetnames:
                ws = wb["Categories"]
                count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[1] and row[1] != "Aucune catégorie":  # name
                        try:
                            db.execute_update("""
                                INSERT OR IGNORE INTO categories (name, name_ar, description)
                                VALUES (?, ?, ?)
                            """, (row[1], row[2] or '', row[3] or ''))
                            count += 1
                        except:
                            pass
                imported_counts["Categories"] = count
            
            # 5. Importer Raccourcis (NEW)
            if "Raccourcis" in wb.sheetnames:
                ws = wb["Raccourcis"]
                count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[1] and row[1] != "Aucun raccourci":  # label
                        try:
                            db.execute_update("""
                                INSERT OR IGNORE INTO pos_shortcuts (label, product_id, category_id, price, image_path, position)
                                VALUES (?, ?, ?, ?, ?, ?)
                            """, (row[1], row[2] or None, row[3] or None, row[4] or None, row[5] or '', row[6] or 0))
                            count += 1
                        except:
                            pass
                imported_counts["Raccourcis"] = count
            summary = "\n".join([f"• {k}: {v} enregistrements" for k, v in imported_counts.items()])
            logger.info(f"Restauration depuis: {filename}")
            QMessageBox.information(self, _('title_success'), _('msg_import_success').format(summary))
            
        except Exception as e:
            logger.error(f"Erreur import excel: {e}")
            QMessageBox.critical(self, _('title_error'), f"Échec: {e}")

    def reset_all_data(self):
        """Réinitialiser toutes les données de l'application"""
        if not auth_manager.check_permission('manage_reset'):
            QMessageBox.warning(self, "Accès refusé", "Vous n'avez pas la permission de réinitialiser l'application.")
            return

        _ = i18n_manager.get
        # Première confirmation
        reply1 = QMessageBox.critical(self, _('group_reset'), 
            _('msg_confirm_reset_1'),
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        if reply1 != QMessageBox.Yes:
            return
        
        # Deuxième confirmation avec mot de passe
        from PyQt5.QtWidgets import QInputDialog
        password, ok = QInputDialog.getText(self, _('title_password_check'),
            _('msg_password_check'),
            QLineEdit.Password)
        
        if not ok or not password:
            return
        
        # Vérifier le mot de passe
        current_user = auth_manager.get_current_user()
        if not current_user:
            QMessageBox.critical(self, _('title_error'), "Utilisateur non connecté")
            return
        
        success, unused_msg, unused_data = auth_manager.login(current_user['username'], password)
        if not success:
            QMessageBox.critical(self, _('title_error'), "Mot de passe incorrect!")
            return
        
        try:
            # Créer une sauvegarde avant suppression
            from datetime import datetime
            from core.backup import backup_manager
            backup_manager.create_backup()
            
            # Supprimer toutes les données
            tables_to_clear = [
                'sale_items',
                'sales',
                'return_items',
                'returns',
                'customer_credit_transactions',
                'supplier_transactions',
                'price_history',
                'products',
                'customers',
                'suppliers',
                'categories',
                'pos_shortcuts',
                'audit_log'
            ]
            
            for table in tables_to_clear:
                try:
                    db.execute_update(f"DELETE FROM {table}")
                except Exception as e:
                    logger.error(f"Erreur suppression {table}: {e}")
            
            logger.info("⚠️ RÉINITIALISATION COMPLÈTE effectuée par l'utilisateur")
            QMessageBox.information(self, _('title_success'), _('msg_reset_success'))
            
        except Exception as e:
            logger.error(f"Erreur réinitialisation: {e}")
            QMessageBox.critical(self, _('title_error'), f"Erreur: {e}")

    def create_users_tab(self):
        """Onglet de gestion des utilisateurs"""
        _ = i18n_manager.get
        tab = QWidget()
        layout = QHBoxLayout()
        
        # Liste des utilisateurs (Gauche)
        list_layout = QVBoxLayout()
        list_layout.addWidget(QLabel(_('label_user_list')))
        
        self.users_table = QTableWidget()
        self.users_table.setColumnCount(4)  # 4 colonnes
        # Headers from i18n
        self.users_table.setHorizontalHeaderLabels(_('table_headers_users'))
        self.users_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.users_table.setColumnWidth(3, 100)
        list_layout.addWidget(self.users_table)
        
        refresh_btn = QPushButton(_('btn_refresh'))
        refresh_btn.clicked.connect(self.load_users)
        list_layout.addWidget(refresh_btn)
        
        layout.addLayout(list_layout, 2)
        
        # Formulaire d'ajout (Droite)
        form_group = QGroupBox(_('group_add_user'))
        form_layout = QFormLayout()
        
        self.username_input = QLineEdit()
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)
        self.fullname_input = QLineEdit()
        self.role_input = QComboBox()
        self.role_input.addItem(_('role_cashier'), "cashier")
        self.role_input.addItem(_('role_admin'), "admin")
        
        form_layout.addRow(_('label_username'), self.username_input)
        form_layout.addRow(_('label_password'), self.password_input)
        form_layout.addRow(_('label_fullname_user'), self.fullname_input)
        form_layout.addRow(_('label_role'), self.role_input)
        
        add_btn = QPushButton(_('btn_create_user'))
        add_btn.setStyleSheet("background-color: #27ae60; color: white; padding: 10px; font-weight: bold;")
        add_btn.clicked.connect(self.add_user)
        form_layout.addRow(add_btn)
        
        form_group.setLayout(form_layout)
        layout.addWidget(form_group, 1)
        
        tab.setLayout(layout)
        return tab

    def create_store_tab(self):
        """Onglet infos magasin"""
        _ = i18n_manager.get
        tab = QWidget()
        layout = QFormLayout()
        
        store_config = config.STORE_CONFIG
        
        self.store_name = QLineEdit(store_config.get('name', ''))
        self.store_phone = QLineEdit(store_config.get('phone', ''))
        self.store_address = QLineEdit(store_config.get('address', ''))
        self.store_city = QLineEdit(store_config.get('city', ''))
        
        self.store_nif = QLineEdit(store_config.get('tax_id', ''))  
        self.store_nis = QLineEdit(store_config.get('nis', ''))      
        self.store_rc = QLineEdit(store_config.get('rc', ''))        
        self.store_ai = QLineEdit(store_config.get('ai', ''))        
        
        save_btn = QPushButton(_('btn_save_store'))
        save_btn.setDefault(True)
        save_btn.setAutoDefault(True)
        save_btn.clicked.connect(self.save_store_settings)
        
        layout.addRow(_('label_store_name'), self.store_name)
        layout.addRow(_('label_store_phone'), self.store_phone)
        layout.addRow(_('label_store_address'), self.store_address)
        layout.addRow(_('label_store_city'), self.store_city)
        layout.addRow(_('label_store_nif'), self.store_nif)
        layout.addRow(_('label_store_nis'), self.store_nis)
        layout.addRow(_('label_store_nis'), self.store_nis)
        layout.addRow(_('label_store_rc'), self.store_rc)
        layout.addRow(_('label_store_ai'), self.store_ai)
        
        # Expiry Warning Days
        self.expiry_days_spin = QSpinBox()
        self.expiry_days_spin.setRange(1, 365)
        self.expiry_days_spin.setSuffix(" " + _('suffix_days'))
        # Load current setting
        expiry_res = db.fetch_one("SELECT setting_value FROM settings WHERE setting_key = 'expiry_warning_days'")
        current_expiry = int(expiry_res['setting_value']) if expiry_res else 7
        self.expiry_days_spin.setValue(current_expiry)
        
        layout.addRow(_('label_expiry_days'), self.expiry_days_spin)
        
        layout.addRow(save_btn)
        
        tab.setLayout(layout)
        return tab

    def save_store_settings(self):
        """Sauvegarder les paramètres du magasin"""
        _ = i18n_manager.get
        try:
            # Update memory
            config.STORE_CONFIG['name'] = self.store_name.text()
            config.STORE_CONFIG['phone'] = self.store_phone.text()
            config.STORE_CONFIG['address'] = self.store_address.text()
            config.STORE_CONFIG['city'] = self.store_city.text()
            config.STORE_CONFIG['tax_id'] = self.store_nif.text()
            config.STORE_CONFIG['nis'] = self.store_nis.text()
            config.STORE_CONFIG['rc'] = self.store_rc.text()
            config.STORE_CONFIG['ai'] = self.store_ai.text()
            
            # Save to DB
            settings_to_save = {
                'store_name': config.STORE_CONFIG['name'],
                'store_phone': config.STORE_CONFIG['phone'],
                'store_address': config.STORE_CONFIG['address'],
                'store_city': config.STORE_CONFIG['city'],
                'store_nif': config.STORE_CONFIG['tax_id'],
                'store_nis': config.STORE_CONFIG['nis'],
                'store_rc': config.STORE_CONFIG['rc'],
                'store_rc': config.STORE_CONFIG['rc'],
                'store_ai': config.STORE_CONFIG['ai'],
                'expiry_warning_days': str(self.expiry_days_spin.value())
            }
            
            for key, value in settings_to_save.items():
                check = db.fetch_one("SELECT id FROM settings WHERE setting_key = ?", (key,))
                if check:
                    db.execute_update("UPDATE settings SET setting_value = ? WHERE setting_key = ?", (value, key))
                else:
                    db.execute_update("INSERT INTO settings (setting_key, setting_value) VALUES (?, ?)", (key, value))
            
            QMessageBox.information(self, _('title_success'), _('msg_store_saved'))
            
        except Exception as e:
            logger.error(f"Erreur sauvegarde store settings: {e}")
            QMessageBox.critical(self, _('title_error'), f"Échec: {e}")
        
    def load_users(self):
        """Charger la liste des utilisateurs"""
        # Fix: Check constraints before accessing table
        if not auth_manager.is_admin():
            return
        if not hasattr(self, 'users_table'):
            return
            
        try:
            # Récupérer id, username, full_name, role, et is_active
            query = "SELECT id, username, full_name, role, is_active FROM users WHERE is_active = 1"
            users = db.execute_query(query)
            
            self.users_table.setRowCount(0)
            for user in users:
                row = self.users_table.rowCount()
                self.users_table.insertRow(row)
                self.users_table.setItem(row, 0, QTableWidgetItem(user['username']))
                self.users_table.setItem(row, 1, QTableWidgetItem(user['full_name']))
                self.users_table.setItem(row, 2, QTableWidgetItem(user['role']))
                
                # Boutons d'action
                btn_widget = QWidget()
                layout = QHBoxLayout(btn_widget)
                layout.setContentsMargins(0, 0, 0, 0)
                layout.setAlignment(Qt.AlignCenter)
                
                # Bouton Modifier mot de passe
                pwd_btn = QPushButton("🔑")
                pwd_btn.setFixedSize(30, 30)
                pwd_btn.setToolTip("Modifier le mot de passe")
                pwd_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #e8f5e9; 
                        border: none; 
                        border-radius: 4px;
                        color: #27ae60;
                    }
                    QPushButton:hover {
                        background-color: #c8e6c9;
                    }
                """)
                pwd_btn.clicked.connect(lambda checked, uid=user['id'], uname=user['username']: self.change_password_dialog(uid, uname))
                
                # Bouton Permissions (Nouveau)
                perm_btn = QPushButton("🛡️")
                perm_btn.setFixedSize(30, 30)
                perm_btn.setToolTip("Gérer les permissions")
                perm_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #e3f2fd; 
                        border: none; 
                        border-radius: 4px;
                        color: #1976d2;
                    }
                    QPushButton:hover {
                        background-color: #bbdefb;
                    }
                """)
                perm_btn.clicked.connect(lambda checked, uid=user['id'], uname=user['username'], urole=user['role']: self.open_permission_dialog(uid, uname, urole))

                del_btn = QPushButton("🗑️")
                del_btn.setFixedSize(30, 30)
                del_btn.setToolTip("Supprimer l'utilisateur")
                del_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #ffcccc; 
                        border: none; 
                        border-radius: 4px;
                        color: #c0392b;
                    }
                    QPushButton:hover {
                        background-color: #ffb3b3;
                    }
                """)
                # Connecter avec l'ID et le Username
                del_btn.clicked.connect(lambda checked, uid=user['id'], uname=user['username']: self.delete_user(uid, uname))
                
                layout.addWidget(pwd_btn)
                layout.addWidget(perm_btn)
                layout.addWidget(del_btn)
                self.users_table.setCellWidget(row, 3, btn_widget)
                
        except Exception as e:
            logger.error(f"Erreur chargement utilisateurs: {e}")

    def open_permission_dialog(self, user_id, username, role):
        """Ouvrir le dialogue de permissions"""
        dialog = PermissionDialog(user_id, username, role, self)
        dialog.exec_()
        
    def change_password_dialog(self, user_id, username):
        """Ouvrir le dialogue pour changer le mot de passe"""
        from PyQt5.QtWidgets import QDialog, QFormLayout
        
        current_user = auth_manager.get_current_user()
        is_own_account = current_user and current_user['id'] == user_id
        
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Modifier le mot de passe - {username}")
        dialog.setMinimumWidth(350)
        
        layout = QVBoxLayout()
        form = QFormLayout()
        
        # Si c'est son propre compte, demander l'ancien mot de passe
        old_pwd_input = None
        if is_own_account:
            old_pwd_input = QLineEdit()
            old_pwd_input.setEchoMode(QLineEdit.Password)
            old_pwd_input.setPlaceholderText("Votre mot de passe actuel")
            form.addRow("Ancien mot de passe:", old_pwd_input)
        
        new_pwd_input = QLineEdit()
        new_pwd_input.setEchoMode(QLineEdit.Password)
        new_pwd_input.setPlaceholderText("Minimum 4 caractères")
        
        confirm_pwd_input = QLineEdit()
        confirm_pwd_input.setEchoMode(QLineEdit.Password)
        
        form.addRow("Nouveau mot de passe:", new_pwd_input)
        form.addRow("Confirmer:", confirm_pwd_input)
        
        layout.addLayout(form)
        
        # Boutons
        btn_layout = QHBoxLayout()
        cancel_btn = QPushButton("Annuler")
        cancel_btn.clicked.connect(dialog.reject)
        
        save_btn = QPushButton("💾 Enregistrer")
        save_btn.setDefault(True)
        save_btn.setAutoDefault(True)
        save_btn.setStyleSheet("background-color: #27ae60; color: white;")
        
        def save_password():
            new_pwd = new_pwd_input.text()
            confirm_pwd = confirm_pwd_input.text()
            
            if new_pwd != confirm_pwd:
                QMessageBox.warning(dialog, "Erreur", "Les mots de passe ne correspondent pas")
                return
            
            if len(new_pwd) < 4:
                QMessageBox.warning(dialog, "Erreur", "Le mot de passe doit contenir au moins 4 caractères")
                return
            
            if is_own_account:
                # Utiliser change_password avec vérification de l'ancien mot de passe
                old_pwd = old_pwd_input.text()
                success, msg = auth_manager.change_password(user_id, old_pwd, new_pwd)
            else:
                # Admin peut réinitialiser sans ancien mot de passe
                from core.security import hash_password
                new_hash = hash_password(new_pwd)
                query = "UPDATE users SET password_hash = ? WHERE id = ?"
                db.execute_update(query, (new_hash, user_id))
                success, msg = True, "Mot de passe modifié avec succès"
            
            if success:
                QMessageBox.information(dialog, "Succès", msg)
                dialog.accept()
            else:
                QMessageBox.warning(dialog, "Erreur", msg)
        
        save_btn.clicked.connect(save_password)
        
        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(save_btn)
        layout.addLayout(btn_layout)
        
        dialog.setLayout(layout)
        dialog.exec_()

    def delete_user(self, user_id, username):
        """Supprimer un utilisateur"""
        current_user = auth_manager.get_current_user()
        
        # 1. Empêcher la suppression de soi-même
        if current_user and current_user['id'] == user_id:
            QMessageBox.warning(self, "Action interdite", "Vous ne pouvez pas supprimer votre propre compte !")
            return

        reply = QMessageBox.question(self, "Confirmation", 
                                   f"Voulez-vous vraiment supprimer l'utilisateur '{username}' ?",
                                   QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                                   
        if reply == QMessageBox.Yes:
            try:
                # 2. Vérifier si c'est un admin et si c'est le dernier
                all_users = auth_manager.get_all_users()
                target_user = next((u for u in all_users if u['id'] == user_id), None)
                
                if target_user and target_user['role'] == 'admin':
                    admin_count = sum(1 for u in all_users if u['role'] == 'admin')
                    if admin_count <= 1:
                        QMessageBox.warning(self, "Erreur", "Impossible de supprimer le dernier administrateur.")
                        return

                # 3. Supprimer (Soft Delete)
                query = "UPDATE users SET is_active = 0 WHERE id = ?"
                rows = db.execute_update(query, (user_id,))
                
                if rows > 0:
                    logger.info(f"Utilisateur supprimé: {username} (ID: {user_id})")
                    QMessageBox.information(self, "Succès", f"L'utilisateur {username} a été supprimé.")
                    self.load_users()
                else:
                    QMessageBox.warning(self, "Erreur", "L'utilisateur n'a pas pu être supprimé.")
                
            except Exception as e:
                logger.error(f"Erreur suppression utilisateur: {e}")
                QMessageBox.critical(self, "Erreur", f"Erreur technique: {str(e)}")
            
    def add_user(self):
        """Ajouter un utilisateur"""
        username = self.username_input.text().strip()
        password = self.password_input.text().strip()
        fullname = self.fullname_input.text().strip()
        role = self.role_input.currentData()
        
        if not all([username, password, fullname]):
            QMessageBox.warning(self, "Erreur", "Tous les champs sont requis")
            return
            
        success, msg, _ = auth_manager.create_user(username, password, fullname, role)
        
        if success:
            QMessageBox.information(self, "Succès", msg)
            self.username_input.clear()
            self.password_input.clear()
            self.fullname_input.clear()
            self.load_users()
        else:
            QMessageBox.critical(self, "Erreur", msg)
            
    def toggle_dark_mode(self):
        """Basculer le mode sombre"""
        is_dark = self.dark_mode_cb.isChecked()
        self.theme_changed.emit(is_dark) # Émettre le signal pour MainWindow

    def create_tutorial_tab(self):
        """Onglet tutoriel d'utilisation"""
        _ = i18n_manager.get
        from PyQt5.QtWidgets import QTextBrowser
        
        tab = QWidget()
        layout = QVBoxLayout()
        
        tutorial = QTextBrowser()
        tutorial.setOpenExternalLinks(True)
        tutorial.setHtml(_("tutorial_content"))
        
        layout.addWidget(tutorial)
        tab.setLayout(layout)
        return tab

    def create_about_tab(self):
        """Onglet à propos"""
        tab = QWidget()
        layout = QVBoxLayout()
        layout.setAlignment(Qt.AlignCenter)
        
        # Logo/Icône
        logo = QLabel()
        logo.setAlignment(Qt.AlignCenter)
        logo_path = str(config.LOGO_PATH)
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path)
            scaled_pixmap = pixmap.scaled(100, 100, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            logo.setPixmap(scaled_pixmap)
        else:
            logo.setText("🏪")
            logo.setStyleSheet("font-size: 72px;")
            
        layout.addWidget(logo)
        
        # Titre
        title = QLabel("DamDev POS")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #667eea;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        # Version
        version = QLabel("Version 1.0.0")
        version.setStyleSheet("font-size: 14px; color: #666;")
        version.setAlignment(Qt.AlignCenter)
        layout.addWidget(version)
        
        layout.addSpacing(30)
        
        # Développeur
        dev_group = QGroupBox("👨‍💻 Développé par")
        dev_group.setStyleSheet("QGroupBox { font-size: 14px; font-weight: bold; }")
        dev_layout = QVBoxLayout()
        
        dev_name = QLabel("DamDev")
        dev_name.setStyleSheet("font-size: 20px; font-weight: bold; color: #667eea;")
        dev_name.setAlignment(Qt.AlignCenter)
        dev_layout.addWidget(dev_name)
        
        dev_layout.addSpacing(10)
        
        phone = QLabel("📞 0561491987")
        phone.setStyleSheet("font-size: 14px;")
        phone.setAlignment(Qt.AlignCenter)
        dev_layout.addWidget(phone)
        
        email = QLabel("📧 amine.drif2002@gmail.com")
        email.setStyleSheet("font-size: 14px;")
        email.setAlignment(Qt.AlignCenter)
        dev_layout.addWidget(email)
        
        dev_group.setLayout(dev_layout)
        layout.addWidget(dev_group)
        
        layout.addSpacing(20)
        
        # Copyright
        copyright_lbl = QLabel("© 2026- Tous droits réservés")
        copyright_lbl.setStyleSheet("color: #999;")
        copyright_lbl.setAlignment(Qt.AlignCenter)
        layout.addWidget(copyright_lbl)
        
        layout.addStretch()
        tab.setLayout(layout)
        return tab
    
    def set_dark_mode(self, is_dark):
        """Appliquer le mode sombre/clair"""
        if is_dark:
            # Mode sombre
            tab_style = """
                QTabWidget::pane {
                    background-color: #2c3e50;
                    border: 1px solid #4a6785;
                    border-radius: 8px;
                }
                QTabBar::tab {
                    background-color: #34495e;
                    color: #ecf0f1;
                    padding: 10px 20px;
                    border: 1px solid #4a6785;
                    border-radius: 5px 5px 0 0;
                }
                QTabBar::tab:selected {
                    background-color: #3498db;
                    color: white;
                }
            """
            group_style = """
                QGroupBox {
                    background-color: #34495e;
                    color: #ecf0f1;
                    border: 1px solid #4a6785;
                    border-radius: 10px;
                    padding: 15px;
                    font-weight: bold;
                }
                QGroupBox::title {
                    color: #ecf0f1;
                    subcontrol-origin: margin;
                    left: 15px;
                }
            """
            table_style = """
                QTableWidget {
                    background-color: #34495e;
                    color: #ecf0f1;
                    gridline-color: #4a6785;
                    border: 1px solid #4a6785;
                    border-radius: 8px;
                }
                QTableWidget::item {
                    padding: 8px;
                    border-bottom: 1px solid #4a6785;
                }
                QTableWidget::item:selected {
                    background-color: #3498db;
                    color: white;
                }
                QTableWidget::item:alternate {
                    background-color: #2c3e50;
                }
                QHeaderView::section {
                    background-color: #3498db;
                    color: white;
                    padding: 10px;
                    border: none;
                    font-weight: bold;
                }
            """
            label_style = "color: #ecf0f1;"
        else:
            # Mode clair
            tab_style = """
                QTabWidget::pane {
                    background-color: #f5f5f5;
                    border: 1px solid #e0e0e0;
                    border-radius: 8px;
                }
                QTabBar::tab {
                    background-color: white;
                    color: #2c3e50;
                    padding: 10px 20px;
                    border: 1px solid #e0e0e0;
                    border-radius: 5px 5px 0 0;
                }
                QTabBar::tab:selected {
                    background-color: #3498db;
                    color: white;
                }
            """
            group_style = """
                QGroupBox {
                    background-color: white;
                    color: #2c3e50;
                    border: 1px solid #e0e0e0;
                    border-radius: 10px;
                    padding: 15px;
                    font-weight: bold;
                }
                QGroupBox::title {
                    color: #2c3e50;
                    subcontrol-origin: margin;
                    left: 15px;
                }
            """
            table_style = """
                QTableWidget {
                    background-color: white;
                    color: #2c3e50;
                    gridline-color: #e0e0e0;
                    border: 1px solid #e0e0e0;
                    border-radius: 8px;
                }
                QTableWidget::item {
                    padding: 8px;
                    border-bottom: 1px solid #e0e0e0;
                }
                QTableWidget::item:selected {
                    background-color: #3498db;
                    color: white;
                }
                QTableWidget::item:alternate {
                    background-color: #f8f9fa;
                }
                QHeaderView::section {
                    background-color: #3498db;
                    color: white;
                    padding: 10px;
                    border: none;
                    font-weight: bold;
                }
            """
            label_style = "color: #2c3e50;"
        
        # Appliquer aux tabs
        if hasattr(self, 'tabs'):
            self.tabs.setStyleSheet(tab_style)
        
        # Appliquer aux tables
        if hasattr(self, 'users_table'):
            self.users_table.setStyleSheet(table_style)
        
        # Appliquer les groupes de façon récursive
        from PyQt5.QtWidgets import QGroupBox
        for group in self.findChildren(QGroupBox):
            group.setStyleSheet(group_style)

