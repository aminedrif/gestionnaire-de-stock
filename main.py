# -*- coding: utf-8 -*-
"""
Application de Gestion de Mini-Market
Point d'entrée principal avec interface PyQt5
"""
import sys
from pathlib import Path

# Ajouter le dossier parent au chemin Python
sys.path.append(str(Path(__file__).parent))

import config
from core.logger import logger
from database.db_manager import db

# Import PyQt5
from PyQt5.QtWidgets import QApplication, QMessageBox
from PyQt5.QtCore import Qt
from ui.login_dialog import LoginDialog
from ui.main_window import MainWindow

def initialize_application():
    """Initialiser l'application"""
    try:
        logger.info("=" * 60)
        logger.info(f"Démarrage de {config.APP_NAME} v{config.APP_VERSION}")
        logger.info("=" * 60)
        
        # Vérifier la base de données
        logger.info("Vérification de la base de données...")
        db_info = db.get_database_info()
        logger.info(f"Base de données: {db_info['path']}")
        logger.info(f"Taille: {db_info['size_bytes'] / 1024:.2f} KB")
        logger.info(f"Tables: {len(db_info['tables'])}")
        
        # Afficher les compteurs de tables
        for table, count in db_info['table_counts'].items():
            logger.info(f"  - {table}: {count} enregistrement(s)")
            

        logger.info("Application initialisée avec succès")
        return True
        
    except Exception as e:
        logger.critical(f"Erreur lors de l'initialisation: {e}")
        logger.exception("Détails de l'erreur:")
        return False

def load_settings_from_db():
    """Charger les paramètres depuis la base de données et mettre à jour config.py"""
    try:
        # Mapping DB keys -> Config keys
        mapping = {
            'store_name': ('store', 'name'),
            'store_address': ('store', 'address'),
            'store_city': ('store', 'city'),
            'store_phone': ('store', 'phone'),
            'store_email': ('store', 'email'),
            'store_nif': ('store', 'tax_id'),
            'store_nis': ('store', 'nis'),
            'store_rc': ('store', 'rc'),
            'store_ai': ('store', 'ai'),
            'tax_rate': ('store', 'tax_rate'),
            'currency': ('store', 'currency'),
        }
        
        # Fetch all settings
        rows = db.execute_query("SELECT setting_key, setting_value FROM settings")
        settings_map = {row['setting_key']: row['setting_value'] for row in rows}
        
        for db_key, (section, config_key) in mapping.items():
            if db_key in settings_map and settings_map[db_key]:
                # Update config
                if section == 'store':
                    config.STORE_CONFIG[config_key] = settings_map[db_key]
                    
        logger.info("Configuration chargée depuis la base de données")
        
    except Exception as e:
        logger.error(f"Erreur chargement configuration: {e}")

def main():
    """Fonction principale avec interface PyQt5"""
    # Initialiser l'application
    # Enable High DPI Scaling
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    
    if not initialize_application():
        print("Erreur lors de l'initialisation de l'application")
        sys.exit(1)
    
    # Charger la configuration depuis la base de données
    load_settings_from_db()
    
    # Créer l'application Qt
    app = QApplication(sys.argv)
    app.setApplicationName(config.APP_NAME)
    app.setApplicationVersion(config.APP_VERSION)
    
    # Configurer le style
    app.setStyle('Fusion')
    
    # Configurer l'icône de l'application (Barre des tâches + Fenêtres)
    import os
    import ctypes
    
    # 1. Définir AppUserModelID pour Windows (pour que l'icône s'affiche dans la barre des tâches)
    myappid = f'damdev.pos.{config.APP_VERSION}' # chaîne arbitraire
    try:
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
    except Exception:
        pass
        
    # 2. Définir l'icône globale
    from PyQt5.QtGui import QIcon
    # 2. Définir l'icône globale
    from PyQt5.QtGui import QIcon
    icon_path = str(config.LOGO_PATH)
    if os.path.exists(icon_path):
        app_icon = QIcon(icon_path)
        app.setWindowIcon(app_icon)
    
    # Vérification de la licence
    from core.license import license_manager
    is_licensed, license_msg = license_manager.is_licensed()
    
    if not is_licensed:
        logger.warning(f"Licence non valide: {license_msg}")
        from ui.license_dialog import LicenseDialog
        license_dialog = LicenseDialog()
        if license_dialog.exec_() != LicenseDialog.Accepted:
            logger.info("Activation de licence annulée - Fermeture de l'application")
            sys.exit(0)
        logger.info("Licence activée avec succès")
    else:
        logger.info(f"Licence valide: {license_msg}")
    
    # Boucle principale de l'application
    while True:
        # Afficher le dialogue de connexion
        login_dialog = LoginDialog()
        
        if login_dialog.exec_() == LoginDialog.Accepted:
            # Connexion réussie, obtenir les données utilisateur
            from core.auth import auth_manager
            user_data = auth_manager.get_current_user()
            
            if user_data:
                # Créer et afficher la fenêtre principale
                main_window = MainWindow(user_data)
                main_window.showMaximized()
                
                # Configurer la sauvegarde automatique
                from PyQt5.QtCore import QTimer
                from core.backup import backup_manager
                
                # Fetch backup configuration from DB
                try:
                    res_enabled = db.fetch_one("SELECT setting_value FROM settings WHERE setting_key = 'auto_backup_enabled'")
                    backup_enabled = (res_enabled['setting_value'] == '1') if res_enabled else config.BACKUP_CONFIG.get("auto_backup", True)
                    
                    res_interval = db.fetch_one("SELECT setting_value FROM settings WHERE setting_key = 'backup_interval_hours'")
                    backup_interval_hours = int(res_interval['setting_value']) if res_interval else config.BACKUP_CONFIG.get("backup_interval_hours", 5)
                except Exception as e:
                    logger.error(f"Error loading backup config: {e}")
                    backup_enabled = True
                    backup_interval_hours = 5

                def perform_auto_backup():
                    """Effectuer une sauvegarde automatique"""
                    try:
                        # Sauvegarde SQL (base de données)
                        backup_manager.auto_backup()
                        
                        # Sauvegarde Excel
                        import openpyxl
                        from datetime import datetime
                        backup_dir = config.BACKUP_DIR
                        backup_dir.mkdir(exist_ok=True)
                        
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        excel_path = backup_dir / f"auto_backup_{timestamp}.xlsx"
                        
                        wb = openpyxl.Workbook()
                        
                        # Produits
                        ws = wb.active
                        ws.title = "Produits"
                        products = db.execute_query("SELECT * FROM products WHERE is_active = 1")
                        if products:
                            ws.append(list(dict(products[0]).keys()))
                            for p in products:
                                ws.append(list(dict(p).values()))
                        
                        # Ventes
                        ws_sales = wb.create_sheet("Ventes")
                        sales = db.execute_query("SELECT * FROM sales")
                        if sales:
                            ws_sales.append(list(dict(sales[0]).keys()))
                            for s in sales:
                                ws_sales.append(list(dict(s).values()))
                        
                        # Clients
                        ws_cust = wb.create_sheet("Clients")
                        customers = db.execute_query("SELECT * FROM customers WHERE is_active = 1")
                        if customers:
                            ws_cust.append(list(dict(customers[0]).keys()))
                            for c in customers:
                                ws_cust.append(list(dict(c).values()))
                        
                        wb.save(excel_path)
                        logger.info(f"Sauvegarde automatique créée: {excel_path.name}")
                        
                    except Exception as e:
                        logger.error(f"Erreur sauvegarde automatique: {e}")
                
                # Timer start if enabled
                if backup_enabled:
                    backup_interval_ms = backup_interval_hours * 3600 * 1000
                    logger.info(f"Auto-backup enabled. Interval: {backup_interval_hours} hours ({backup_interval_ms} ms)")
                    backup_timer = QTimer()
                    backup_timer.timeout.connect(perform_auto_backup)
                    backup_timer.start(backup_interval_ms)
                else:
                    logger.info("Auto-backup is disabled.")
                
                # Sauvegarde à la fermeture de l'application
                app.aboutToQuit.connect(perform_auto_backup)
                
                # Lancer la boucle d'événements
                app.exec_()
                
                # Si on arrive ici, c'est que la fenêtre principale a été fermée
                # Si l'utilisateur est toujours connecté, c'est une fermeture normale -> Quitter
                # Si l'utilisateur est déconnecté, c'est un logout -> Boucler
                if auth_manager.is_authenticated():
                    break
            else:
                logger.error("Erreur: Aucune donnée utilisateur après connexion")
                QMessageBox.critical(None, "Erreur", "Erreur lors de la récupération des données utilisateur")
                sys.exit(1)
        else:
            # Connexion annulée
            logger.info("Connexion annulée par l'utilisateur")
            sys.exit(0)

if __name__ == "__main__":
    main()
